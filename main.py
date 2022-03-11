from openpyxl import load_workbook
from datetime import timedelta, datetime
from kivy.app import App
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.core.window import Window
import pickle

Window.size = (360, 640)


def cal_average(dic):
    """average tape time calculation from database"""
    total_t = timedelta(minutes=0)
    total_c = sum(map(int, dic.keys()))
    for t in dic.values():
        time_d = timedelta(hours=t.hour, minutes=t.minute,
                           seconds=t.second)
        total_t += time_d
    one_turn = total_t / total_c
    return one_turn


now = datetime.today().strftime('%Y-%m-%d')  # Current date
with open("settings.bin", "rb") as f:
    settings = pickle.load(f)
if settings['model'] is not None:
    wb = load_workbook('data/database.xlsx')
    base_models = str(len(wb.sheetnames))
    ws = wb[settings['model']]
    tapemeter = dict()
    for row in ws.iter_rows(min_row=4, max_col=2, values_only=True):
        tapemeter[row[0]] = row[1]
    count_one = cal_average(tapemeter)
    current_base = settings['model']
else:
    base_models = 0
    count_one = timedelta(0, 0, 0)
    current_base = "Выбрать базу данных"


class HomePage(Screen):
    Screen.current_base = current_base
    Screen.base_models = base_models
    Screen.count_one = str(count_one.total_seconds()) + ' сек.'

    def login_btn_press(self):
        self.ids.login_img1.source = 'img/login_2.png'

    def login_btn_rel(self):
        self.ids.login_img1.source = 'img/login_1.png'

    def calculator(self, value, count=count_one):
        if value and count > timedelta(0, 0, 0):
            tape_len = (count_one * int(value))\
                       * 2 - timedelta(minutes=0, seconds=20)
            # get rid from microseconds:
            tape_side = tape_len / 2
            tape_len = tape_len - timedelta(
                microseconds=tape_len.microseconds)
            tape_side = tape_side - timedelta(
                microseconds=tape_side.microseconds)
            self.display_cass.text = str(tape_len)
            self.display_side.text = str(tape_side)
        else:
            self.ids.delete_left.color = (1, 0, 0, 1)
            self.ids.delete_right.color = (1, 0, 0, 1)


class DataPage(Screen):
    pass


class SettingsPage(Screen):
    pass


class DataScroll(Screen):
    pass


class PageManager(ScreenManager):
    pass


gui = Builder.load_file("kvcode.kv")


class TapeApp(App):
    def build(self):
        return gui


if __name__ == "__main__":
    TapeApp().run()
