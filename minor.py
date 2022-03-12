from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment
from datetime import timedelta, datetime
import pickle
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from googleapiclient.discovery import build
import pprint
from openpyxl.styles import Font
import kivy

now = datetime.today().strftime('%Y-%m-%d')  # Current date
with open("data/settings.bin", "rb") as f:
    settings = pickle.load(f)
wb = load_workbook('data/database.xlsx')
current_base = settings['model']
base_models = str(len(wb.sheetnames))
ws = wb[settings['model']]
tapemeter = dict()


def cal_average(dic):
    """average tape time calculation from database"""
    total_t = timedelta(minutes=0)
    total_c = sum(map(int, dic.keys()))
    for t in dic.values():
        time_d = timedelta(hours=t.hour, minutes=t.minute,
                           seconds=t.second)
        total_t += time_d
    one_turn = total_t / total_c
    return one_turn


def settings_read():
    """create and delete settings for programm"""
    with open("data/settings.bin", "rb") as f:
        s = pickle.load(f)
    print("Настройки программы:")
    act = input("1 - Посмотреть\n"
                "2 - Сбросить \n"
                "0 - Выход\n")
    if act == '1':
        for key, val in s.items():
            print(key, ':', val)
    elif act == '2':
        warn = input("Внимание: сброс насроек!\n"
                     "1 - Продолжить\n"
                     "0 - Выход\n")
        if warn == '1':
            f = open("data/settings.bin", "wb")
            s = {'first_name': None,
                 'last_name': None,
                 'company': 'Grimy Can',
                 'model': "A&D GX-Z5300",
                 'API_key': 'AIzaSyA1XOqp_WF778aez3b0WQI9TxLloOsWBQ8',
                 'database_folder': '1VJoUPOPJeSAMEC6gnx_Jo3EHDTUIHP2',
                 'database_size': None,
                 'database_date': None}
            pickle.dump(s, f)
            print("Настройки сброшены.")
        else:
            pass
    else:
        pass


# work with API Google Drive"
def get_drive_dir_info():
    # work with API Google Drive
    # http://datalytics.ru/all/rabotaem-s-api-google-drive-s-pomoschyu-python/"
    pp = pprint.PrettyPrinter(indent=4)
    SCOPES = ['https://www.googleapis.com/auth/drive']
    SERVICE_ACCOUNT_FILE = 'key.json'
    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    service = build('drive', 'v3', credentials=credentials)
    results = service.files().list(pageSize=10,
                                   fields="nextPageToken, "
                                          "files("
                                          "id, name, mimeType)").execute()

    pp.pprint(results)


def create_database(model):
    wbook = load_workbook('data/database.xlsx')
    wsheet = wbook.create_sheet(model)
    wsheet.sheet_properties.tabColor = "0000FF00"
    # wbook.save('data/database.xlsx')
    wsheet.merge_cells('A1:C2')
    wsheet['A1'].comment = Comment(f"База создана {now}", 'Tapemeter')
    wsheet["A1"].style = '20 % - Accent5'
    wsheet["A3"] = "Счётчик:"
    wsheet["B3"] = "Время:"
    wsheet["C3"] = "Дата:"
    wsheet["A3"].style = '20 % - Accent1'
    wsheet["B3"].style = '20 % - Accent1'
    wsheet["C3"].style = '20 % - Accent1'
    wbook.save('data/database.xlsx')


# open database file with stored data and get one turn:


def add_new_data(counter, time):
    today = datetime.today().strftime('%Y-%m-%d')
    read_count = int(counter)
    read_time_d = timedelta(
        hours=time[0],
        minutes=time[1],
        seconds=time[2])
    # write new data to file:
    ws.cell(ws.max_row, 1).value = read_count
    ws.cell(ws.max_row, 2).value = read_time_d
    ws.cell(ws.max_row, 3).value = today
    wb.save('data/database.xlsx')

print(kivy.__version__)
